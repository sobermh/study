"""
@author:maohui
@time:2022/5/24 10:14
"""

from openpyxl import load_workbook
from openpyxl import workbook

def huoqu():
    wb1=load_workbook('C:/Users/PC/Desktop/1.xlsx')

    sheet_lc=wb1['临床信息']
    sheet_zh=wb1['质荷比']
    for cell_row in sheet_lc['C2':'K158']:
        list1=[]
        for cell in cell_row:
            list1.append(cell.value)
        listID.append(list1[0])
        listsexy.append(list1[3])
        listage.append(list1[4])
        listmsg.append(list1[5])
        listnum.append(list1[-1])
    for cell_row in sheet_zh['B4':'H603']:
        list2=[]
        for cell in cell_row:
            list2.append(cell.value)
        print(list2)
        if list2[1]!='误差（ppm）':
            if list2[1]=='均值':
                list2[1]='——'
            listpic.append(list2[1])
            listzhihebi_j.append(list2[2])
            listzhihebi_k.append(list2[3])
            listzhihebi_l.append(list2[4])
            listzhihebi_m.append(list2[5])
            listzhihebi_n.append(list2[6])

def xieru(listID,listsexy,listage,listmsg,listpic,listzhihebi_j,listzhihebi_k,listzhihebi_l,listzhihebi_m,listzhihebi_n):
    wb2=load_workbook("C:/Users/PC/Desktop/2.xlsx")
    sheetIn=wb2['Sheet2']
    #处理有些单元格合并，有些未合并(行，列，值)
    #num
    for i in range(1,len(listID)+1):
        flag=5+(i-1)*4
        sheetIn.cell(flag, 1, i)
        sheetIn.merge_cells(start_row=flag, start_column=1, end_row=flag+3, end_column=1)
    #id
    for i in range(0,len(listID)):
        flag=5+i*4
        sheetIn.cell(flag, 2, listID[i - 1])
        sheetIn.merge_cells(start_row=flag, start_column=2, end_row=flag + 3, end_column=2)
    #sexy
    for i in range(0,len(listsexy)):
        flag=5+i*4
        sheetIn.cell(flag, 3, listsexy[i - 1])
        sheetIn.merge_cells(start_row=flag, start_column=3, end_row=flag + 3, end_column=3)
    #age
    for i in range(0, len(listage)):
        flag = 5 + i * 4
        sheetIn.cell(flag, 4, listage[i - 1])
        sheetIn.merge_cells(start_row=flag, start_column=4, end_row=flag + 3, end_column=4)
    #msg
    for i in range(0,len(listmsg)):
        flag=5+i*4
        sheetIn.cell(flag, 8, listmsg[i - 1])
        sheetIn.merge_cells(start_row=flag, start_column=8, end_row=flag + 3, end_column=8)

    # pic
    for i in range(1,len(listpic)+1):
        sheetIn.cell(i+4, 9, listpic[i - 1])
    #,listzhihebi_j,listzhihebi_k,listzhihebi_l,listzhihebi_m,listzhihebi_n
    for i in range(0,len(listzhihebi_j)):
            sheetIn.cell(i+5, 10, listzhihebi_j[i])
    for i in range(0,len(listzhihebi_k)):
            sheetIn.cell(i+5, 11, listzhihebi_k[i])
    for i in range(0,len(listzhihebi_l)):
            sheetIn.cell(i+5, 12, listzhihebi_l[i])
    for i in range(0,len(listzhihebi_m)):
            sheetIn.cell(i+5, 13, listzhihebi_m[i])
    for i in range(0,len(listzhihebi_n)):
            sheetIn.cell(i+5, 14, listzhihebi_n[i])
    # 受试者排除标准、中止试验标准
    for i in range(5, sheetIn.max_row + 1):
        sheetIn.cell(i, 6, '/')
    for i in range(5, sheetIn.max_row + 1):
        sheetIn.cell(i, 7, '/')
    wb2.save("C:/Users/PC/Desktop/2.xlsx")

if __name__=="__main__":
    listID=[]
    listsexy=[]
    listage=[]
    listmsg=[]
    listnum=[]
    #
    listpic=[]
    listzhihebi_j=[]
    listzhihebi_k = []
    listzhihebi_l = []
    listzhihebi_m = []
    listzhihebi_n = []
    huoqu()
    try:
        xieru(listID,listsexy,listage,listmsg,listpic,listzhihebi_j,listzhihebi_k,listzhihebi_l,listzhihebi_m,listzhihebi_n)
        print('succeed')
    except:
        print('shibai')