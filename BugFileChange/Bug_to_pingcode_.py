"""
@author:maohui
@time:2022/3/2 9:31
"""


from openpyxl import load_workbook
from openpyxl import workbook


#读取原生的bug_excel文件
wb1=load_workbook('C:/Users/PC/Desktop/GZH-09-01.xlsx')
#激活要操作的表
sheet1=wb1['Sheet1']
#获取需要导出的单元格的列最大值
maxcell=str(sheet1.max_row)
maxcellmodel1='B'+maxcell
maxcelltype1='C'+maxcell
maxcelldesc1='F'+maxcell
maxcellstep1='G'+maxcell
maxcellplan1='J'+maxcell
#获取导出bug_excel的连续单元格
cellmodel1=sheet1['B12':maxcellmodel1]
celltype1=sheet1['C12':maxcelltype1]
celldesc1=sheet1['F12':maxcelldesc1]
cellstep1=sheet1['G12':maxcellstep1]
cellplan1=sheet1['J12':maxcellplan1]
#获取bug_excel的功能模块内容
listmodel=[]
for i in cellmodel1:
    for row in i:
        listmodel.append(row.value)
#获取bug_excel的缺陷类别内容
listtype=[]
for i in celltype1:
    for row in i:
        listtype.append(row.value)
#获取bug_excel的缺陷概述内容
listdesc=[]
for i in celldesc1:
    for row in i:
        listdesc.append(row.value)
#获取bug_excel的复现步骤内容
liststep=[]
for i in cellstep1:
    for row in i:
        liststep.append(row.value)
#获取bug_excel的解决方案内容
listplan=[]
for i in cellplan1:
    for row in i:
        listplan.append(row.value)
#处理获取到的数据
listtitle=[]
for index in range(len(cellmodel1)):
    listtitle1=listmodel[index]+'---'+listdesc[index]
    listtitle.append(listtitle1)


#读取要输入的excel的文件
wb2=load_workbook('C:/Users/PC/Desktop/PingCode.Agile-defects-import-template1.xlsx')
#激活要输入excel的表
sheet2=wb2['defect']
#获取输入表的单元格的列最大值(使用输出表bug_excel的)
maxcell2=str((sheet1.max_row)-9)
maxcelltitle2='A'+maxcell2
maxcellstep2='L'+maxcell2
maxcelltype2='Q'+maxcell2
maxcellplan2='I'+maxcell2
#获取输入表的连续单元格
celltitle2=sheet2['A3':maxcelltitle2]
cellstep2=sheet2['L3':maxcellstep2]
celltype2=sheet2['Q3':maxcelltype2]
cellplan2=sheet2['I3':maxcellplan2]
#获取导入excel表标题的单元格,并导入内容
listtitle2=[]
for r in celltitle2:
    for i in r:
        listtitle2.append(i)
for i in range(len(listtitle2)):
    cell=listtitle2[i]
    cell.value=listtitle[i]
#获取导入excel表重现步骤的单元格,并导入内容
liststep2=[]
for r in cellstep2:
    for i in r:
        liststep2.append(i)
for i in range(len(liststep2)):
    cell=liststep2[i]
    cell.value=liststep[i]
#获取导入excel表缺陷类型的单元格,并导入内容
listtype2=[]
for r in celltype2:
    for i in r:
        listtype2.append(i)
for i in range(len(listtype2)):
    cell=listtype2[i]
    cell.value=listtype[i]
#获取导入excel表解决方案的单元格,并导入内容
listplan2=[]
for r in cellplan2:
    for i in r:
        listplan2.append(i)
for i in range(len(listplan2)):
    cell=listplan2[i]
    cell.value=listplan[i]
wb2.save('C:/Users/PC/Desktop/PingCode.Agile-defects-import-template1.xlsx')