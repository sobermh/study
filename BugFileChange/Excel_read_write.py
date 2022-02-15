"""
@author:maohui
@time:2022/2/10 14:45
"""
from openpyxl import workbook
from openpyxl import load_workbook

# 读取文件excel文件

wb = load_workbook('C:/Users/PC/Desktop/1.xlsx')
wb2 = load_workbook('C:/Users/PC/Desktop/bug.xlsx')
# #取正在活跃的表单
# ws = wb.active()

# 激活sheet1表
sheet = wb['sheet1']
sheet2=wb2['Sheet1']
# #1.获取单元格的值
# cell=sheet['A1']
# print(cell.value)
# #2.获取单元格的值
# cell1=sheet.cell(row=1,column=1)
# print(cell1.value,cell.row,cell.column,cell.coordinate)

# 获取导出最大单元格
maxcelltitle='B'+str(sheet.max_row)
maxcellnum='A'+str(sheet.max_row)
maxcelltype='G'+str(sheet.max_row)
maxcellmodel='C'+str(sheet.max_row)
maxCellLevel= 'F'+str(sheet.max_row)

#导入表最大的单元格行
maxcelltitle2='F'+str(sheet.max_row+10)
maxcellnum2='A'+str(sheet.max_row+10)
maxcelltype2='C'+str(sheet.max_row+10)
maxcellmodel2='B'+str(sheet.max_row+10)
maxCellLevel2='E'+str(sheet.max_row+10)

# 获取导出excel表标题连续单元格
cellstitle = sheet['B2':maxcelltitle]
#获取导出excel表编号连续单元格
cellsnum=sheet['A2':maxcellnum]
#获取导出excel表缺陷类别连续单元格
cellstype=sheet['G2':maxcelltype]
#获取导出excel表缺陷类型连续单元格
cellsLevel=sheet['F2':maxCellLevel]

# 获取导入excel表标题连续单元格
cellstitle2 = sheet2['F12':maxcelltitle2]
#获取导入excel表编号连续单元格
cellsnum2=sheet2['A12':maxcellnum2]
#获取导入excel表缺陷类别连续单元格
cellstype2=sheet2['C12':maxcelltype2]
#获取导入excel表缺陷模块的连续单元格
cellsmodel2=sheet2['B12':maxcellmodel2]
#获取导入excel表缺陷模块的连续单元格
cellsLevel2=sheet2['E12':maxCellLevel2]

#必须要是这种循环
#获取导出excel表的标题内容
listtitle=[]
for i in cellstitle:
    for r in i:
        listtitle.append(r.value)
#获取导出excel表的编号内容
listnum=[]
for i in cellsnum:
    for r in i:
        listnum.append(r.value)
#获取导出excel表的缺陷类型内容
listtype=[]
for i in cellstype:
    for r in i:
        listtype.append(r.value)
#获取导出excel表严重程度的内容
listLevel=[]
for i in cellsLevel:
    for r in i:
        listLevel.append(r.value)
#处理获取的标题内容
listmodel=[]
listtitle1=[]
for i in listtitle:
    i1=str(i).split('---')
    listmodel.append(i1[0])
    listtitle1.append(i1[-1])

#获取导入excel表标题的单元格
list2=[]
for r in cellstitle2:
    for i in r:
            list2.append(i)
#将导出excel表的标题值赋予导入excel表的标题单元格中
for i in range(len(list2)):
        cell=list2[i]
        cell.value=listtitle1[i]
#获取导入excel表缺陷模块的单元格
list5=[]
for r in cellsmodel2:
    for i in r:
        list5.append(i)
#将导出excel表的标题值赋予导入excel表的缺陷模块单元格中
for i in range(len(list5)):
        cell=list5[i]
        cell.value=listmodel[i]

#获取导入excel表编号的单元格
list3=[]
for r in cellsnum2:
    for i in r:
            list3.append(i)
#将导出excel表的编号值赋予导入excel表的标题单元格中
for i in range(len(list3)):
        cell=list3[i]
        cell.value=listnum[i]
#获取导入excel表缺陷类型的单元格
list4=[]
for r in cellstype2:
    for i in r:
            list4.append(i)
#将导出excel表的缺陷类型值赋予导入excel表的缺陷类型单元格中
for i in range(len(list4)):
        cell=list4[i]
        cell.value=listtype[i]
#获取导入excel表严重程度的单元格
list7=[]
for r in cellsLevel2:
    for i in r:
            list7.append(i)
#将导出excel表的严重程度值赋予导入excel表的缺陷类型单元格中
for i in range(len(list7)):
        cell=list7[i]
        cell.value=listLevel[i]
wb2.save('C:/Users/PC/Desktop/bug.xlsx')
# #获取最大行号
# print(sheet.max_row)
# #获取最大列号
# print(sheet.max_column)

# 改变单元格的值
# ws['B1'].value="test"
# 保存wb，改变才能生效
# wb.save('C:/Users/PC/Desktop/1.xlsx')

# 打印单元格的值
# print(ws['B1'].value)

# 创建不同的一个新表
# wb.create_sheet("test")

# 打印所有的表
# print(wb.sheetnames)

# 得到所有的工作表
# sheets=wb.get_sheet_names()
# print(sheets)
