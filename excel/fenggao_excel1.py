"""
@author:maohui
@time:2022/5/24 10:14
"""

from openpyxl import load_workbook
from openpyxl import workbook
import xlsxwriter

def huoqu():
    wb1=load_workbook('1.xlsx', data_only=True)

    sheet_lc=wb1['临床信息']
    sheet_zh=wb1['峰高']
    for cell_row in sheet_lc['C2':'K158']:
        list1=[]
        for cell in cell_row:
            list1.append(cell.value)
        listID.append(list1[0])
        listsexy.append(list1[3])
        listage.append(list1[4])
        listmsg.append(list1[5])
        listnum.append(list1[-1])
    for cell_row in sheet_zh['B4':'H483']:
        list2=[]
        for cell in cell_row:
            list2.append(cell.value)
        print(list2)
        if list2[1]!='误差（ppm）':
            if list2[1]=='均值':
                list2[1]='——'
            listpic.append(list2[1])
            listzhihebi_j.append(list2[2])
            listzhihebi_k.append(list2[3])
            listzhihebi_l.append(list2[4])
            listzhihebi_m.append(list2[5])
            listzhihebi_n.append(list2[6])


def xieru(listID,listsexy,listage,listmsg,listpic,listzhihebi_j,listzhihebi_k,listzhihebi_l,listzhihebi_m,listzhihebi_n):
    workbook = xlsxwriter.Workbook('峰高.xlsx')
    center={
        'align': 'center',  # 水平居中对齐
        'valign': 'vcenter',  # 垂直居中对齐'
        'text_wrap': 1,
        'border': 5,  # 边框宽度
    }
    sheet1 = {
        'align': 'center',  # 水平居中对齐
        'valign': 'vcenter',  # 垂直居中对齐'
        'text_wrap': 1,
        'border': 5,  # 边框宽度
        'font_size': 14
    }
    sheet_2={
        'align': 'center',  # 水平居中对齐
        'valign': 'vcenter',  # 垂直居中对齐'
        'text_wrap': 1,
        'border': 5,  # 边框宽度
        'font':"Wingdings 2",######使用R可以插入单选框
        'font_size': 14
    }
    sheet_title={
        'align': 'center',  # 水平居中对齐
        'valign': 'vcenter',  # 垂直居中对齐'
        'text_wrap': 1,
        'bold': True,  # 是否粗体
        'border': 5,  # 边框宽度
    }
    worksheet = workbook.add_worksheet('Sheet1')


#字体样式
    superscript = workbook.add_format({'font_script': 1})
    style1=workbook.add_format(sheet1)
    style2=workbook.add_format(sheet_2)
    style_title=workbook.add_format(sheet_title)
    style_center=workbook.add_format(center)

    #title
    worksheet.merge_range('A2:O2',"飞行时间质谱仪 临床研究原始记录A",style_title)
    worksheet.merge_range('A3:O3', "试验日期   2022  年 5 月 5 日     入组采样日期   2022  年 5 月 5 日", style_title)
    worksheet.merge_range('A4:A5',"序号",style_title)
    worksheet.merge_range('B4:B5', "受试者ID", style_title)
    worksheet.merge_range('C4:C5', "性别", style_title)
    worksheet.merge_range('D4:D5', "年龄", style_title)
    worksheet.merge_range('E4:E5', "样本入选标准", style_title)
    worksheet.merge_range('F4:F5', "受试者排除标准", style_title)
    worksheet.merge_range('G4:G5', "中止试验标准", style_title)
    worksheet.merge_range('H4:H5', "临床诊断", style_title)
    worksheet.merge_range('I4:I5', "质谱图编号", style_title)
    worksheet.merge_range('J4:N4', "峰强度", style_title)
    # worksheet.merge_range('F4:F5', "序号", style_title)
    worksheet.write('J5', 926.35, style_title)
    worksheet.write('K5', 2933.03, style_title)
    worksheet.write('L5', 3973.46, style_title)
    worksheet.write('M5', 5905.23, style_title)
    worksheet.write('N5', 9282.82, style_title)
    worksheet.merge_range('O4:O5', "偏移≤1000ppm", style_title)
    #处理有些单元格合并，有些未合并(行，列，值)
    #num
    for i in range(1,len(listID)+1):
        flag=6+(i-1)*4
        worksheet.merge_range('A%d:A%d'%(flag,flag+3),"%03d"%i,style1)
    #id
    for i in range(1,len(listID)+1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('B%d:B%d' % (flag, flag + 3),listID[i-1],style1)
    #sexy
    for i in range(1,len(listsexy)+1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('C%d:C%d' % (flag, flag + 3), listsexy[i - 1],style1)
    #age
    for i in range(1, len(listage)+1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('D%d:D%d' % (flag, flag + 3), listage[i - 1],style1)
    #msg
    for i in range(1,len(listmsg)+1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('H%d:H%d' % (flag, flag + 3), listmsg[i - 1],style1)
    # pic
    for i in range(1,len(listpic)+1):
        worksheet.write_column(5,8,listpic,style1)
    #,listzhihebi_j,listzhihebi_k,listzhihebi_l,listzhihebi_m,listzhihebi_n
    count = 1
    for i in range(0,len(listzhihebi_j)):
        if count != 4:
            worksheet.write(i + 5, 9,listzhihebi_j[i],style1)
        else:
            if type(listzhihebi_j[i])==float:
                worksheet.write_rich_string("J%d" % (i + 6),
                                            superscript, "平均值",
                                            str(round(listzhihebi_j[i])), style1)
            else:
                worksheet.write_rich_string("J%d"%(i+6),
                                        superscript,"平均值",
                                        str(listzhihebi_j[i]),style1)
            count=0
        count+=1
    for i in range(0,len(listzhihebi_k)):
        if count != 4:
            worksheet.write(i + 5, 10, listzhihebi_k[i],style1)
        else:
            if type(listzhihebi_k[i])==float:
                worksheet.write_rich_string("K%d" % (i + 6),
                                            superscript, "平均值",
                                            str(round(listzhihebi_k[i])), style1)
            else:
                worksheet.write_rich_string("K%d" % (i + 6),
                                        superscript, "平均值",
                                        str(listzhihebi_k[i]),style1)
            count = 0
        count += 1
    for i in range(0,len(listzhihebi_l)):
        if count != 4:
            worksheet.write(i + 5, 11, listzhihebi_l[i],style1)
        else:
            if type(listzhihebi_l[i])==float:
                worksheet.write_rich_string("L%d" % (i + 6),
                                            superscript, "平均值",
                                            str(round(listzhihebi_l[i])), style1)
            else:
                worksheet.write_rich_string("L%d" % (i + 6),
                                        superscript, "平均值",
                                        str(listzhihebi_l[i]),style1)
            count = 0
        count += 1

    for i in range(0,len(listzhihebi_m)):
        if count != 4:
            worksheet.write(i + 5, 12, listzhihebi_m[i],style1)
        else:
            if type(listzhihebi_m[i]) == float:
                worksheet.write_rich_string("M%d" % (i + 6),
                                            superscript, "平均值",
                                            str(round(listzhihebi_m[i])), style1)
            else:
                worksheet.write_rich_string("M%d" % (i + 6),
                                        superscript, "平均值",
                                        str(listzhihebi_m[i]),style1)
            count = 0
        count += 1
    for i in range(0,len(listzhihebi_n)):
        if count != 4:
            worksheet.write(i + 5, 13, listzhihebi_n[i],style1)
        else:
            if type(listzhihebi_n[i]) == float:
                worksheet.write_rich_string("N%d" % (i + 6),
                                            superscript, "平均值",
                                            str(round(listzhihebi_n[i])), style1)
            else:
                worksheet.write_rich_string("N%d" % (i + 6),
                                        superscript, "平均值",
                                        str(listzhihebi_n[i]),style1)
            count = 0
        count += 1

    # 受试者排除标准、中止试验标准
    for i in range(1, len(listmsg)+1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('F%d:F%d' % (flag, flag + 3), '/',style1)
    for i in range(1, len(listmsg) + 1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('G%d:G%d' % (flag, flag + 3), '/',style1)
    # 样本入选标准
    for i in range(1, len(listmsg) + 1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('E%d:E%d' % (flag, flag + 3),1,style1)
        worksheet.write_rich_string("E%d" % (flag),
                                     style1, "是 ",
                                     style2, "S",
                                     style1,"\n否 □",
                                    style_center)
    # 偏移
    for i in range(1, len(listmsg) + 1):
        flag = 6 + (i - 1) * 4
        worksheet.merge_range('O%d:O%d' % (flag, flag + 3), 1, style1)
        worksheet.write_rich_string("O%d" % (flag),
                                    style1, "是 ",
                                    style2, "S",
                                    style1, "\n否 □",
                                    style_center)
    #设置列宽(字符)
    worksheet.set_column('A:A',3)
    worksheet.set_column('C:D',4)
    worksheet.set_column('J:N', 10)
    worksheet.set_column('H:H', 12)
    workbook.close()




if __name__=="__main__":
    listID=[]
    listsexy=[]
    listage=[]
    listmsg=[]
    listnum=[]
    #
    listpic=[]
    listzhihebi_j=[]
    listzhihebi_k = []
    listzhihebi_l = []
    listzhihebi_m = []
    listzhihebi_n = []
    huoqu()
    # try:
    xieru(listID,listsexy,listage,listmsg,listpic,listzhihebi_j,listzhihebi_k,listzhihebi_l,listzhihebi_m,listzhihebi_n)
    print('succeed')

    # except:
    #     print('shibai')
